from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parent
UNPROCESSED_DIR = ROOT / "未处理"
BAR_HEIGHT = 237
TITLE_COLOR = "#9C6A5A"
VALUE_COLOR = "#D94841"
SEPARATOR_COLOR = "#E8DCC2"
DEFAULT_BG = (255, 255, 240)
TITLE_SIZE = 15
VALUE_SIZE = 23
TITLE_FONT_CANDIDATES = [
    Path(r"C:\Windows\Fonts\msyh.ttc"),
    Path(r"C:\Windows\Fonts\simhei.ttf"),
]
VALUE_FONT_CANDIDATES = [
    Path(r"C:\Windows\Fonts\msyhbd.ttc"),
    Path(r"C:\Windows\Fonts\simhei.ttf"),
]

FIELD_ORDER = [
    "年限",
    "品种",
    "交易机会类型",
    "时间级别",
    "机会释放复杂度",
    "积累内部结构复杂度",
    "释放形式(正向/逆向)",
    "边线是否共振",
    "释放复杂度",
    "是否到达目标位",
    "最大盈利倍数",
    "加仓位置",
    "加仓释放复杂度",
    "加仓最大盈利倍数",
    "数据来源",
    "相关因子",
    "目标位之后是否延续释放",
    "编号",
]

DESTINATION_MAP = {
    ("上通道", "H1"): Path("通道") / "上通道" / "H1",
    ("上通道", "H4"): Path("通道") / "上通道" / "H4",
    ("下通道", "H1"): Path("通道") / "下通道" / "H1",
    ("下通道", "H4"): Path("通道") / "下通道" / "H4",
    ("双底", "H1"): Path("双头双底") / "双底" / "H1",
    ("双底", "H4"): Path("双头双底") / "双底" / "H4",
    ("双头", "H1"): Path("双头双底") / "双头" / "H1",
    ("双头", "H4"): Path("双头双底") / "双头" / "H4",
    ("复合头肩顶", "H1"): Path("头肩形态") / "复合头肩顶" / "H1",
    ("复合头肩顶", "H4"): Path("头肩形态") / "复合头肩顶" / "H4",
    ("头肩顶", "H1"): Path("头肩形态") / "头肩顶" / "H1",
    ("头肩顶", "H4"): Path("头肩形态") / "头肩顶" / "H4",
    ("头肩底", "H1"): Path("头肩形态") / "头肩底" / "H1",
    ("头肩底", "H4"): Path("头肩形态") / "头肩底" / "H4",
    ("上旗", "H1"): Path("旗形") / "上旗" / "H1",
    ("上旗", "H4"): Path("旗形") / "上旗" / "H4",
    ("上旗子", "H1"): Path("旗形") / "上旗" / "H1",
    ("上旗子", "H4"): Path("旗形") / "上旗" / "H4",
    ("下旗", "H1"): Path("旗形") / "下旗" / "H1",
    ("下旗", "H4"): Path("旗形") / "下旗" / "H4",
    ("上三", "H1"): Path("三角形") / "上三" / "H1",
    ("上三", "H4"): Path("三角形") / "上三" / "H4",
    ("下三", "H1"): Path("三角形") / "下三" / "H1",
    ("下三", "H4"): Path("三角形") / "下三" / "H4",
}


@dataclass
class CaseRow:
    values: dict[str, str]

    @property
    def case_id(self) -> str:
        return self.values["编号"]

    @property
    def opportunity_type(self) -> str:
        return self.values["交易机会类型"]

    @property
    def timeframe(self) -> str:
        return self.values["时间级别"]


def find_excel_file() -> Path:
    matches = sorted(
        path for path in ROOT.glob("*.xlsx") if not path.name.startswith("~$")
    )
    if not matches:
        raise FileNotFoundError("No Excel workbook found in the case overview directory.")
    return matches[0]


def load_case_rows() -> dict[str, CaseRow]:
    workbook = load_workbook(find_excel_file(), data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    rows: dict[str, CaseRow] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[-1]:
            continue
        values = {
            str(header): "" if value is None else str(value).strip()
            for header, value in zip(headers, row)
        }
        rows[values["编号"]] = CaseRow(values=values)
    return rows


def load_font(candidates: Iterable[Path], size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for path in candidates:
        if path.exists():
            return ImageFont.truetype(str(path), size=size)
    return ImageFont.load_default()


def quantize_color(pixel: tuple[int, int, int]) -> tuple[int, int, int]:
    return tuple(min(255, int(round(channel / 5) * 5)) for channel in pixel)


def detect_background_color(image: Image.Image) -> tuple[int, int, int]:
    rgb = image.convert("RGB")
    width, height = rgb.size
    samples: list[tuple[int, int, int]] = []
    for y in range(10, max(11, height - 10), 20):
        for x in range(10, max(11, width - 10), 20):
            pixel = rgb.getpixel((x, y))
            if sum(pixel) < 690:
                continue
            if max(pixel) - min(pixel) > 45:
                continue
            samples.append(quantize_color(pixel))
    if not samples:
        return DEFAULT_BG
    return Counter(samples).most_common(1)[0][0]


def text_width(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> int:
    if not text:
        return 0
    left, _, right, _ = draw.textbbox((0, 0), text, font=font)
    return right - left


def fit_single_line(
    draw: ImageDraw.ImageDraw,
    text: str,
    font_paths: Iterable[Path],
    initial_size: int,
    max_width: int,
    min_size: int,
) -> tuple[ImageFont.ImageFont, str]:
    font_size = initial_size
    while font_size >= min_size:
        font = load_font(font_paths, font_size)
        if text_width(draw, text, font) <= max_width:
            return font, text
        font_size -= 1

    font = load_font(font_paths, min_size)
    truncated = text
    while truncated and text_width(draw, f"{truncated}...", font) > max_width:
        truncated = truncated[:-1]
    return font, f"{truncated}..." if truncated and truncated != text else truncated


def wrap_text(
    draw: ImageDraw.ImageDraw,
    text: str,
    font_paths: Iterable[Path],
    initial_size: int,
    max_width: int,
    max_lines: int,
    min_size: int,
) -> tuple[ImageFont.ImageFont, list[str]]:
    for font_size in range(initial_size, min_size - 1, -1):
        font = load_font(font_paths, font_size)
        lines: list[str] = []
        current = ""
        for char in text:
            test = f"{current}{char}"
            if current and text_width(draw, test, font) > max_width:
                lines.append(current)
                current = char
            else:
                current = test
        if current:
            lines.append(current)

        if len(lines) <= max_lines:
            return font, lines

    font = load_font(font_paths, min_size)
    lines = []
    current = ""
    for char in text:
        test = f"{current}{char}"
        if current and text_width(draw, test, font) > max_width:
            lines.append(current)
            current = char
        else:
            current = test
    if current:
        lines.append(current)

    clipped = lines[:max_lines]
    if len(lines) > max_lines and clipped:
        last = clipped[-1]
        while last and text_width(draw, f"{last}...", font) > max_width:
            last = last[:-1]
        clipped[-1] = f"{last}..." if last else "..."
    return font, clipped


def render_case_image(source: Path, row: CaseRow, destination: Path) -> None:
    with Image.open(source) as original:
        chart = original.convert("RGB")
        width, height = chart.size
        background = detect_background_color(chart)
        canvas = Image.new("RGB", (width, height + BAR_HEIGHT), background)
        canvas.paste(chart, (0, 0))

    draw = ImageDraw.Draw(canvas)
    title_font = load_font(TITLE_FONT_CANDIDATES, TITLE_SIZE)
    value_font = load_font(VALUE_FONT_CANDIDATES, VALUE_SIZE)

    bar_top = height
    draw.line((0, bar_top + 3, width, bar_top + 3), fill=SEPARATOR_COLOR, width=2)

    outer_padding = 16
    cell_gap = 8
    inner_padding = 4
    usable_width = width - outer_padding * 2
    cell_width = (usable_width - cell_gap * 8) / 9
    row_title_y = [bar_top + 18, bar_top + 123]
    row_value_y = [bar_top + 46, bar_top + 151]

    for index, field in enumerate(FIELD_ORDER):
        row_index = 0 if index < 9 else 1
        column = index % 9
        x = outer_padding + column * (cell_width + cell_gap)
        title_max_width = int(cell_width - inner_padding * 2)
        title_font_fit, title_text = fit_single_line(
            draw=draw,
            text=field,
            font_paths=TITLE_FONT_CANDIDATES,
            initial_size=TITLE_SIZE,
            max_width=title_max_width,
            min_size=12,
        )
        draw.text(
            (x + inner_padding, row_title_y[row_index]),
            title_text,
            font=title_font_fit,
            fill=TITLE_COLOR,
        )

        value = row.values.get(field, "")
        value_font_fit, lines = wrap_text(
            draw=draw,
            text=value,
            font_paths=VALUE_FONT_CANDIDATES,
            initial_size=VALUE_SIZE,
            max_width=title_max_width,
            max_lines=2,
            min_size=17,
        )
        line_height = value_font_fit.size + 4 if hasattr(value_font_fit, "size") else 24
        for line_number, line in enumerate(lines):
            draw.text(
                (x + inner_padding, row_value_y[row_index] + line_number * line_height),
                line,
                font=value_font_fit,
                fill=VALUE_COLOR,
            )

    destination.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(destination)


def destination_for(row: CaseRow) -> Path:
    key = (row.opportunity_type, row.timeframe)
    if key not in DESTINATION_MAP:
        raise KeyError(f"No destination rule for {key[0]} + {key[1]}.")
    return ROOT / DESTINATION_MAP[key] / f"{row.case_id}.png"


def main() -> None:
    case_rows = load_case_rows()
    sources = sorted(
        path for path in UNPROCESSED_DIR.glob("EUR*.png") if path.is_file()
    )
    if not sources:
        print("No matching images found in 未处理.")
        return

    processed = []
    skipped = []
    for source in sources:
        case_id = source.stem
        if case_id not in case_rows:
            skipped.append((case_id, "Excel 中不存在该编号"))
            continue
        row = case_rows[case_id]
        try:
            destination = destination_for(row)
        except KeyError as exc:
            skipped.append((case_id, str(exc)))
            continue
        render_case_image(source=source, row=row, destination=destination)
        source.unlink()
        processed.append((case_id, destination.relative_to(ROOT).as_posix()))

    for case_id, destination in processed:
        print(f"{case_id} -> {destination}")
    for case_id, reason in skipped:
        print(f"{case_id} skipped: {reason}")


if __name__ == "__main__":
    main()
